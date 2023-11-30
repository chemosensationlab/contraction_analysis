# -*- coding: utf-8 -*-
"""
Created on Tue May  9 09:55:25 2023

@author: wiesbrock
"""


import pandas as pd
import numpy as np
import matplotlib.pylab as plt
import seaborn as sns
from scipy.stats import zscore
import numpy as np
from scipy.signal import argrelextrema
import math
import os

os.chdir(r'C:\Users\wiesbrock\Desktop\Projekte\Uterus\Uterus')




def find_sequences(arr):
    # Entferne alle Nullen aus dem Array
    arr = np.diff(arr, 1)
    arr[arr==2]=1
    arr[arr==-2]=-1
    arr[(arr==2)+1]=1
    arr[(arr==-2)+1]=-1
    arr = [x for x in arr if x != 0]
    sequences = []
    sequence_end=[]

    # Überprüfe, ob die Länge des Arrays weniger als 4 beträgt (keine vollständigen Sequenzen möglich)
    if len(arr) < 4:
        return sequences,sequence_end

    # Iteriere über das Array und suche nach Sequenzen
    sequences = []
    sequence_end=[]
    if len(arr)>4:
    
        last_seq_end = 0  # Speichert das Ende der letzten gefundenen Sequenz
        for i in range(len(arr) - 3):
            # Überprüfen, ob die Sequenz mindestens Länge 4 hat und sich nicht innerhalb der letzten gefundenen Sequenz befindet
            #if i >= last_seq_end and (arr[i:i+4] == [1, -1, -1, 1] or arr[i:i+4] == [-1, 1, 1, -1]):
                if i >= last_seq_end and (arr[i:i+4] == [1, -1, -1, 1]):
                    #if (arr[i:i+4] == [1, -1, -1, 1]):
                        sequences.append(i)
                        sequence_end.append(i+3)
    return sequences,sequence_end
            
                  # Aktualisiere das Ende der letzten gefundenen Sequenz
                  
def find_sequences_neg(arr):
    # Entferne alle Nullen aus dem Array
    arr = np.diff(arr, 1)
    arr[arr==2]=1
    arr[arr==-2]=-1
    arr[(arr==2)+1]=1
    arr[(arr==-2)+1]=-1
    arr = [x for x in arr if x != 0]
    sequences = []
    sequence_end=[]

    # Überprüfe, ob die Länge des Arrays weniger als 4 beträgt (keine vollständigen Sequenzen möglich)
    if len(arr) < 4:
        return sequences,sequence_end

    # Iteriere über das Array und suche nach Sequenzen
    sequences = []
    sequence_end=[]
    if len(arr)>4:
    
        last_seq_end = 0  # Speichert das Ende der letzten gefundenen Sequenz
        for i in range(len(arr) - 3):
            # Überprüfen, ob die Sequenz mindestens Länge 4 hat und sich nicht innerhalb der letzten gefundenen Sequenz befindet
            #if i >= last_seq_end and (arr[i:i+4] == [1, -1, -1, 1] or arr[i:i+4] == [-1, 1, 1, -1]):
                if i >= last_seq_end and (arr[i:i+4] == [-1, 1, 1, -1]):
                    #if (arr[i:i+4] == [1, -1, -1, 1]):
                        sequences.append(i)
                        sequence_end.append(i+3)
    return sequences,sequence_end
            
                  # Aktualisiere das Ende der letzten gefundenen Sequenz
        
           
            

                                


def set_values_to_zero(trace, points, threshold):
    for p in points:
        
        # Nach links gehen
        i = p
        while i >= 0 and trace[i] > threshold:
            trace[i] = np.nan
            i -= 1
        # Nach rechts gehen
        i = p+1
        while i < len(trace) and trace[i] > threshold:
            trace[i] = np.nan
            i += 1

def find_nth_element(arr, n):
    count = 0
    for i, elem in enumerate(arr):
        if elem == 1 or elem == -1:
            count += 1
            if count == n:
                return i
    # n ist größer als die Anzahl der Elemente in arr, die die Bedingung erfüllen
    return None
    

def binary_conversion(arr,value):
    #binary conversion, use z-normalized array
    binary=np.zeros((len(arr)))
    binary[arr>.35]=value
    binary[arr<-.35]=-value
    
    if binary[0]==1:
        binary[0]=0
        
    if binary[0]==-1:
        binary[0]=0
    
    return binary

def binary_conversion_small(arr,value):
    #binary conversion, use z-normalized array
    binary=np.zeros((len(arr)))
    binary[arr>.25]=value
    binary[arr<-.25]=-value
    if binary[0]==1:
        binary[0]=0
        
    if binary[0]==-1:
        binary[0]=0
    
    return binary

def binary_conversion_tiny(arr,value):
    #binary conversion, use z-normalized array
    binary=np.zeros((len(arr)))
    binary[arr>.15]=value
    binary[arr<-.15]=-value
    if binary[0]==1:
        binary[0]=0
        
    if binary[0]==-1:
        binary[0]=0
    
    return binary

def asym_std(arr):
    
    if not isinstance(arr, np.ndarray):
        raise TypeError("Input must be a NumPy array")
        
    row_positive=[]
    row_positive.append(row[row>0])
    row_positive.append(row[row>0]*-1)
    row_positive.append(row[row==0])
    row_positive.append(row[row==0])
    row_positive=np.concatenate(row_positive)

    row_negative=[]
    row_negative.append(row[row<0])
    row_negative.append(row[row<0]*-1)
    row_negative.append(row[row==0])
    row_negative.append(row[row==0])
    row_negative=np.concatenate(row_negative)


    std_positive=np.std(row_positive)
    std_negative=np.std(row_negative)
    
    return std_positive, std_negative

def get_indices_from_trace(binary,pos):
    indices_from_trace=[]
    pos=np.array(pos)
    indices_from_trace=np.zeros((len(pos)))
    diff_binary=np.diff(binary)
    diff_binary[diff_binary==-2]=-1
    diff_binary[diff_binary==2]=1
    for i in range(len(pos)):
        
        
        indices_from_trace[i]=find_nth_element(diff_binary,pos[i])
    
    
        
    return indices_from_trace

def iterative_registration(arr,threshold):
    #normalized trace
    points=np.where(arr>2)[0]
    
    help_array=np.zeros(arr.shape)
    for p in points:
        
        
        i=0
        
        while (arr[p+i]>threshold).all(axis=None):
            i=i+1
            if (p+i < help_array.size).all():
                help_array[p+i] = 1
            else:
                break
            
        j=0
        while (arr[p-j]>threshold).all(axis=None):
            j=j+1
            
            help_array[p-j]=1
            
    points=np.where(arr<-2)
    for p in points:
        i=0
        while (arr[p+i]<-threshold).all(axis=None):
            i=i+1
            
            if (p+i < help_array.size).all():
                help_array[p+i] = 1
            else:
                break
        j=0
        while (arr[p-j]<-threshold).all(axis=None):
            j=j+1
            help_array[p-j]=1
            
        return help_array
    
def indices_of_peaks(z):
    indices_of_pos_peaks=np.where(z>2)[0]
    indices_of_neg_peaks=np.where(z<-2)[0]
    
    indices_peaks=np.concatenate((indices_of_pos_peaks,indices_of_neg_peaks))
    
    return indices_of_pos_peaks,indices_of_neg_peaks
    
   
import openpyxl
for start,stop in [0,1200],[1200,2400],[2400,3600]:
    # Neue Excel-Datei erstellen
    workbook= openpyxl.Workbook()
    workbook.create_sheet('List',2)
    
    
    # Neues Arbeitsblatt erstellen
    worksheet= workbook.active 
    
    
        
    
    sheet='VU'
    exp='all'
    start=start
    stop=stop
    num=0
    data=pd.read_excel(r"Y:\File transfer\Christopher_transfer\VU\output.xlsx")
    names=data.columns
    threshold_high=60
    threshold_low=15
    start_col = 1
    duration_list=[]
    amp_raw_list=[]
    amp_norm_list=[]
    num_of_events_list=[]
    for m in names:
        print(m)
        row=data[m]
        row[row>1.7]=0
        row[row<-1.7]=0
        #row[row==np.max(row)]=np.nan
        #row[row==np.min(row)]=np.nan
        if len(row)>start:
            
            #row[row==np.max(row)]=np.nan
            #row[row==np.min(row)]=np.nan
            big_events_start=[]
            small_events_start=[]
            tiny_events_start=[]
            big_events_stop=[]
            small_events_stop=[]
            tiny_events_stop=[]
            big_events_start_neg=[]
            small_events_start_neg=[]
            tiny_events_start_neg=[]
            big_events_stop_neg=[]
            small_events_stop_neg=[]
            tiny_events_stop_neg=[]
        
    
        text_name=sheet+'_'+str(m)+str(start)+'_'+str(stop)
        
        kernel_size = 9
        kernel = np.ones(kernel_size) / kernel_size
        row = np.convolve(row, kernel, mode='same')
        
        
        z=np.zeros((len(row)))
    
        z[row>0]=row[row>0]/np.max(row[row>0])
        z[row<0]=row[row<0]/np.min(row[row<0])*-1
        
        z=z[start:stop]
        x=np.linspace(0,len(z),len(z)).astype(int)
            
        
            
    
        binary=binary_conversion(z,1)
        
        first_it_binary=binary
        
    
        kernel_size = 10
        kernel = np.ones(kernel_size) / kernel_size
        binary = np.convolve(binary, kernel, mode='same')
        binary[binary>0]=1
        binary[binary<-0]=-1
        
        
    
        diff_binary=np.diff(binary)
        diff_binary_high=np.where(diff_binary==2)[0]
        diff_binary_low=np.where(diff_binary==-2)[0]
    
        diff_binary_high_plus=np.zeros((len(diff_binary_high)))
        diff_binary_low_plus=np.zeros((len(diff_binary_low)))
    
        diff_binary_high_plus=diff_binary_high_plus.astype(int)
        diff_binary_low_plus=diff_binary_low_plus.astype(int)
            
        for i in range(len(diff_binary_low)):
            diff_binary_low_plus[i]=diff_binary_low[i]+1
            
        for i in range(len(diff_binary_high)):
            diff_binary_high_plus[i]=diff_binary_high[i]+1
        
        binary[diff_binary_low]=0
        binary[diff_binary_high]=0
        binary[diff_binary_low_plus]=0
        binary[diff_binary_high_plus]=0
        
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences(binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
            
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
        
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
        
        big_events=indices_start,indices_end
        
        
        filtered_row=row 
        filtered_binary=binary
        
        for i in range(len(indices_start)):
            filtered_row[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            filtered_binary[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
        
        
        
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences_neg(filtered_binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
            
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
        big_events_neg=indices_start,indices_end
        
        for i in range(len(indices_start)):
            filtered_row[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
        
        filtered_row[filtered_row==np.max(row)]=np.nan
        filtered_row[filtered_row==np.min(row)]=np.nan  
        
        z=np.zeros((len(filtered_row)))
          
        z[filtered_row>0]=filtered_row[filtered_row>0]/np.max(filtered_row[filtered_row>0])
        z[filtered_row<0]=filtered_row[filtered_row<0]/np.min(filtered_row[filtered_row<0])*-1
        
        z=z[start:stop]
            
        binary=binary_conversion_small(z,1)
    
    
        kernel_size = 10
        kernel = np.ones(kernel_size) / kernel_size
        binary = np.convolve(binary, kernel, mode='same')
        binary[binary>0]=1
        binary[binary<-0]=-1
        
        if binary[0]==1 or binary[0]==-1:
            binary[0]=0
    
        diff_binary=np.diff(binary)
        diff_binary_high=np.where(diff_binary==2)[0]
        diff_binary_low=np.where(diff_binary==-2)[0]
    
        diff_binary_high_plus=np.zeros((len(diff_binary_high)))
        diff_binary_low_plus=np.zeros((len(diff_binary_low)))
    
        diff_binary_high_plus=diff_binary_high_plus.astype(int)
        diff_binary_low_plus=diff_binary_low_plus.astype(int)
    
        for i in range(len(diff_binary_low)):
            diff_binary_low_plus[i]=diff_binary_low[i]+1
        
        for i in range(len(diff_binary_high)):
            diff_binary_high_plus[i]=diff_binary_high[i]+1
        
        binary[diff_binary_low]=0
        binary[diff_binary_high]=0
        binary[diff_binary_low_plus]=0
        binary[diff_binary_high_plus]=0
        
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences(binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
    
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
            
        nan_check=np.zeros((len(indices_start))).astype(int)
        
        for i in range(len(nan_check)):
            if len(np.where(np.isnan(row[indices_start[i]-start:indices_end[i]-start])==True)[0])>0:
                nan_check[i]=1
                
        
            
        
        
        small_events=indices_start[nan_check==0],indices_end[nan_check==0]
        
        for i in range(len(indices_start)):
            filtered_row[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            binary[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences_neg(binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
    
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
            
        nan_check=np.zeros((len(indices_start))).astype(int)
        
        for i in range(len(nan_check)):
            if len(np.where(np.isnan(row[indices_start[i]-start:indices_end[i]-start])==True)[0])>0:
                nan_check[i]=1
                
        
            
        
        
        small_events_neg=indices_start[nan_check==0],indices_end[nan_check==0]
        
        for i in range(len(indices_start)):
            filtered_row[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            
        filtered_row[filtered_row==np.max(row)]=np.nan
        filtered_row[filtered_row==np.min(row)]=np.nan  
        z=np.zeros((len(filtered_row)))
          
        z[filtered_row>0]=filtered_row[filtered_row>0]/np.max(filtered_row[filtered_row>0])
        z[filtered_row<0]=filtered_row[filtered_row<0]/np.min(filtered_row[filtered_row<0])*-1
        
        z=z[start:stop]
            
        binary=binary_conversion_tiny(z,1)
    
    
        kernel_size = 10
        kernel = np.ones(kernel_size) / kernel_size
        binary = np.convolve(binary, kernel, mode='same')
        binary[binary>0]=1
        binary[binary<-0]=-1
        
        #binary[first_it_binary!=0]==np.nan
    
        diff_binary=np.diff(binary)
        diff_binary_high=np.where(diff_binary==2)[0]
        diff_binary_low=np.where(diff_binary==-2)[0]
    
        diff_binary_high_plus=np.zeros((len(diff_binary_high)))
        diff_binary_low_plus=np.zeros((len(diff_binary_low)))
    
        diff_binary_high_plus=diff_binary_high_plus.astype(int)
        diff_binary_low_plus=diff_binary_low_plus.astype(int)
    
        for i in range(len(diff_binary_low)):
            diff_binary_low_plus[i]=diff_binary_low[i]+1
        
        for i in range(len(diff_binary_high)):
            diff_binary_high_plus[i]=diff_binary_high[i]+1
        
        binary[diff_binary_low]=0
        binary[diff_binary_high]=0
        binary[diff_binary_low_plus]=0
        binary[diff_binary_high_plus]=0
        
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences(binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
    
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
            
        nan_check=np.zeros((len(indices_start))).astype(int)
        
        for i in range(len(nan_check)):
            if len(np.where(np.isnan(row[indices_start[i]-start:indices_end[i]-start])==True)[0])>0:
                nan_check[i]=1
                
        
            
        
        
        tiny_events=indices_start[nan_check==0],indices_end[nan_check==0]
        
        for i in range(len(indices_start)):
            filtered_row[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            binary[indices_start[i]-start-5:indices_end[i]-start+5]=np.nan
            
        pos_start=[]
        pos_end=[]
        if len(binary[binary!=0])>0:
            pos_start,pos_end=find_sequences_neg(binary)
            pos_start=np.array(pos_start)
            pos_end=np.array(pos_end)
        
        if len(pos_start)>0 and pos_start[0]==0:
            pos_start[0]=1
    
        indices_start=[]
        indices_start=get_indices_from_trace(binary,pos_start)
        indices_start=indices_start.astype(int)+start
    
        indices_end=[]
        indices_end=get_indices_from_trace(binary,pos_end)
        indices_end=indices_end.astype(int)+start
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
            
        nan_check=np.zeros((len(indices_start))).astype(int)
        
        for i in range(len(nan_check)):
            if len(np.where(np.isnan(row[indices_start[i]-start:indices_end[i]-start])==True)[0])>0:
                nan_check[i]=1
                
        
            
        
        
        tiny_events_neg=indices_start[nan_check==0],indices_end[nan_check==0]
        
        
            
        big_events_start.append(big_events[0])
        big_events_stop.append(big_events[1])
        big_events_start_neg.append(big_events_neg[0])
        big_events_stop_neg.append(big_events_neg[1])
        small_events_start.append(small_events[0])
        small_events_stop.append(small_events[1])
        small_events_start_neg.append(small_events_neg[0])
        small_events_stop_neg.append(small_events_neg[1])
        tiny_events_start.append(tiny_events[0])
        tiny_events_stop.append(tiny_events[1])
        tiny_events_start_neg.append(tiny_events_neg[0])
        tiny_events_stop_neg.append(tiny_events_neg[1])
           
        
        worksheet.cell(row=1, column=start_col).value = 'raw_amp '+str(m)
        worksheet.cell(row=1, column=start_col+1).value = 'norm_amp '+str(m)
        #worksheet.cell(row=1, column=start_col+2).value = 'difference'+str(m)
        #worksheet.cell(row=1, column=start_col+3).value = 'amp_sum'+str(m)
        
        
        indices_diff=indices_end-indices_start
        indices_end=indices_end[indices_diff<threshold_high]
        indices_start=indices_start[indices_diff<threshold_high]
        indices_diff=indices_diff[indices_diff<threshold_high]
        indices_end=indices_end[indices_diff>threshold_low]
        indices_start=indices_start[indices_diff>threshold_low]
        indices_diff=indices_diff[indices_diff>threshold_low]
        
        big_events_start=np.concatenate(big_events_start)
        big_events_stop=np.concatenate(big_events_stop)
        small_events_start=np.concatenate(small_events_start)
        small_events_stop=np.concatenate(small_events_stop)
        tiny_events_start=np.concatenate(tiny_events_start)
        tiny_events_stop=np.concatenate(tiny_events_stop)
        big_events_start_neg=np.concatenate(big_events_start_neg)
        big_events_stop_neg=np.concatenate(big_events_stop_neg)
        small_events_start_neg=np.concatenate(small_events_start_neg)
        small_events_stop_neg=np.concatenate(small_events_stop_neg)
        tiny_events_start_neg=np.concatenate(tiny_events_start_neg)
        tiny_events_stop_neg=np.concatenate(tiny_events_stop_neg)
        
            
        events_start=np.concatenate((big_events_start,small_events_start,tiny_events_start,big_events_start_neg,small_events_start_neg,tiny_events_start_neg))
        events_start=np.sort(events_start)
            
        events_stop=np.concatenate((big_events_stop,small_events_stop,tiny_events_stop,big_events_stop_neg,small_events_stop_neg,tiny_events_stop_neg))
        events_stop=np.sort(events_stop)
        
        array1 = events_start
        array2 = events_stop
        
        events_diff=events_stop-events_start
        
        duration_list.append(events_diff)
        
        num_of_events_list=[]
        
        #data=pd.read_excel(r"C:\Users\wiesbrock\Desktop\Uterus\Uterus\Daten Muscle\longitudinal.xlsx", sheet_name=sheet)
        row=data[m]
        row[row>1.7]=0
        row[row<-1.7]=0
        #row[row==np.max(row)]=np.nan
        #row[row==np.min(row)]=np.nan
        x=np.linspace(0,len(row),len(row))
        #plt.figure()
        #plt.title(m)
        #plt.plot(x,row)
        #plt.plot(x[events_start],row[events_start],'ro')
        #plt.plot(x[events_stop],row[events_stop],'ko')
        
        
        amp_high=np.zeros((len(events_start)))
        amp_low=np.zeros((len(events_start)))
        duration_amp_high=np.zeros((len(events_start)))
        for i in range(len(events_start)):
            amp_low[i]=np.min(row[events_start[i]:events_stop[i]])
            amp_high[i]=np.max(row[events_start[i]:events_stop[i]])
            duration_amp_high[i]=len(np.where(row[events_start[i]:events_stop[i]]>0)[0])
            
            
        amp_low=np.absolute(amp_low)
        
        
        amp_sum=amp_low+amp_high
        
        
        
        array3=events_diff/2
        
        if len(amp_sum)>0:
            array4=amp_high
            array2=duration_amp_high/2
        else:
            array4=[]
            array2=[]
        
        amp_raw_list.append(array4)
        amp_norm_list.append(array2)
        
        
        for j in range(len(array1)):
            worksheet.cell(row=j+2, column=start_col).value = array4[j]
            worksheet.cell(row=j+2, column=start_col+1).value = array2[j]
            #worksheet.cell(row=j+2, column=start_col+2).value = array3[j]
            #worksheet.cell(row=j+2, column=start_col+3).value = array4[j]
            
    
        # Inkrementiere den Zähler für die Spaltennummer
        start_col += 2
        
    workbook.save('final_2/V17'+str(sheet)+'_'+str(exp)+'_'+str(start)+'_to_'+str(stop)+".xlsx")
    
    wb = openpyxl.load_workbook('final_2/V17'+str(sheet)+'_'+str(exp)+'_'+str(start)+'_to_'+str(stop)+".xlsx")
    n_sheet = workbook['List']
    
    n_sheet.cell(row=1, column=1).value = 'raw_amp '
    n_sheet.cell(row=1, column=2).value = 'norm_amp '
    
    # Definieren Sie Ihre Liste
    meine_liste = np.array(amp_raw_list)
    meine_liste=np.concatenate(meine_liste)
    
    
    # Fügen Sie Ihre Liste in das zweite Sheet ein
    for row_index, row_value in enumerate(meine_liste):
        n_sheet.cell(row=row_index+2, column=1, value=row_value)
        
    meine_liste = np.array(amp_norm_list)
    meine_liste=np.concatenate(meine_liste)
    
    
    # Fügen Sie Ihre Liste in das zweite Sheet ein
    for row_index, row_value in enumerate(meine_liste):
        n_sheet.cell(row=row_index+2, column=2, value=row_value)
    
    # Speichern Sie die Änderungen
    workbook.save('final_2/V17'+str(sheet)+'_'+str(exp)+'_'+str(start)+'_to_'+str(stop)+".xlsx")
