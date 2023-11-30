import numpy as np
import scipy.stats as stats
import pandas as pd
import openpyxl
import matplotlib.pylab as plt
import glob
import os

path = r'Y:\File transfer\Christopher_transfer\VU\*.xlsx'
file_list = glob.glob(path)

output_file = 'output.xlsx'  # Name der Ausgabedatei

# Maximale Länge der Daten ermitteln
max_length = 0
for file_path in file_list:
    x = pd.read_excel(file_path, sheet_name='x movement (change), smoothed')
    if len(x) > max_length:
        max_length = len(x)

# Ausgabedatei erstellen
workbook = openpyxl.Workbook()
worksheet = workbook.active

startcol = 1
for n in range(len(file_list)):
    file_name = os.path.basename(file_list[n])[:-5]  # Pfad kürzen und die letzten 5 Stellen entfernen
    x = pd.read_excel(file_list[n], sheet_name='x movement (change), smoothed')
    y = pd.read_excel(file_list[n], sheet_name='y movement (change), smoothed')
    x_names = x.columns

    for i in range(len(x_names)):
        x_roi = x[x_names[i]]
        y_roi = y[x_names[i]]
        x_fit = x_roi[0:1200]
        y_fit = y_roi[0:1200]
        x_cumsum = np.cumsum(x_fit)
        y_cumsum = np.cumsum(y_fit)
        slope, intercept, r, p, se = stats.linregress(x_cumsum, y_cumsum)
        a = slope
        L = np.sqrt((a ** 2 + 1))
        x_factor = 1 / L
        y_factor = a / L
        projected_velocity = np.zeros((max_length))
        for j in range(len(x_roi)):
            projected_velocity[j] = x_roi[j] * x_factor + y_roi[j] * y_factor

        # Prüfen, ob die Spalte nicht leer ist
        if np.any(projected_velocity):
            # Neuen Spaltennamen erstellen
            col_name = f"{file_name}_{x_names[i]}"

            # Daten in Excel-Datei schreiben
            worksheet.cell(row=1, column=startcol, value=col_name)
            for j in range(len(projected_velocity)):
                worksheet.cell(row=j + 2, column=startcol, value=projected_velocity[j])
            startcol += 1

# Ausgabedatei speichern
workbook.save(output_file)
